"""
PetroLedger — Daily Summary Report Generator.

Produces a daily PDF + Excel report covering all shifts for a given site
and date.

PDF contents:
  • Date, site name, total revenue breakdown (cash / UPI / card / fleet)
  • Per-shift summary table (S1, S2, S3) with variance status
  • Per-attendant shortage summary
  • Payment-mode breakdown with percentages
  • 7-day trend comparison

Excel contents (one sheet per section):
  • Summary sheet with totals
  • Per-shift detail
  • Transaction-level UPI / POS listing
  • Anomaly flags
"""

from __future__ import annotations

import os
import uuid
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

import structlog
from jinja2 import Environment, select_autoescape
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.fms import (
    CashEntry,
    FleetTransaction,
    FmsTransaction,
    FmsTxnStatus,
    PosBatchSettlement,
)
from app.models.pump import Pump
from app.models.reconciliation import ReconciliationResult
from app.models.shift import Shift
from app.models.transaction import UPITransaction
from app.utils.datetime import to_ist

log = structlog.stdlib.get_logger("petroledger.reports.daily")

_REPORTS_DIR = Path(os.environ.get("REPORTS_DIR", "./reports"))


# ── Jinja2 HTML Template ─────────────────────────────────────────────────

_DAILY_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <style>
    body { font-family: Arial, sans-serif; font-size: 11px; color: #333; margin: 0; padding: 20px; }
    h1 { font-size: 18px; margin-bottom: 2px; }
    h2 { font-size: 13px; margin-top: 16px; margin-bottom: 6px; border-bottom: 1px solid #ccc; padding-bottom: 3px; }
    table { width: 100%; border-collapse: collapse; margin-bottom: 10px; }
    th { background: #1a3a5c; color: white; padding: 6px 8px; text-align: left; }
    td { padding: 5px 8px; border-bottom: 1px solid #eee; }
    tr:nth-child(even) td { background: #f7f7f7; }
    .total-row td { font-weight: bold; background: #eef4ff; }
    .match { color: #16a34a; }
    .shortage { color: #dc2626; }
    .excess { color: #d97706; }
    .right { text-align: right; }
    .stat-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin-bottom: 16px; }
    .stat-box { background: #f0f4ff; border: 1px solid #c7d7f5; border-radius: 4px; padding: 10px; }
    .stat-box .val { font-size: 16px; font-weight: bold; color: #1a3a5c; }
    .stat-box .lbl { font-size: 9px; color: #666; margin-top: 2px; }
    .trend-cell { font-size: 9px; color: #666; }
    .footer { margin-top: 30px; font-size: 9px; color: #999; border-top: 1px solid #ccc; padding-top: 6px; }
  </style>
</head>
<body>

<h1>&#9981; PetroLedger — Daily Summary Report</h1>
<div style="color: #666; font-size: 10px; margin-bottom: 12px;">CONFIDENTIAL — OWNER ONLY</div>

<div style="margin-bottom: 12px;">
  <strong>Site:</strong> {{ site_name }} &nbsp;|&nbsp;
  <strong>Date:</strong> {{ report_date }} &nbsp;|&nbsp;
  <strong>Shifts processed:</strong> {{ shift_count }}
</div>

<div class="stat-grid">
  <div class="stat-box"><div class="val">₹{{ total_fms }}</div><div class="lbl">Total FMS Revenue</div></div>
  <div class="stat-box"><div class="val">₹{{ total_cash }}</div><div class="lbl">Total Cash Collected</div></div>
  <div class="stat-box"><div class="val">₹{{ total_digital }}</div><div class="lbl">Digital Payments</div></div>
  <div class="stat-box">
    <div class="val {{ 'shortage' if net_variance > 0 else ('excess' if net_variance < 0 else 'match') }}">
      {% if net_variance == 0 %}MATCH{% elif net_variance > 0 %}₹{{ net_variance }} SHORT{% else %}₹{{ net_variance|abs }} EXCESS{% endif %}
    </div>
    <div class="lbl">Net Variance</div>
  </div>
</div>

<h2>Payment Mode Breakdown</h2>
<table>
  <tr><th>Mode</th><th class="right">Amount (₹)</th><th class="right">% of Revenue</th></tr>
  <tr><td>Cash</td><td class="right">{{ total_cash }}</td><td class="right">{{ cash_pct }}%</td></tr>
  <tr><td>UPI</td><td class="right">{{ total_upi }}</td><td class="right">{{ upi_pct }}%</td></tr>
  <tr><td>Card / POS</td><td class="right">{{ total_card }}</td><td class="right">{{ card_pct }}%</td></tr>
  <tr><td>Fleet</td><td class="right">{{ total_fleet }}</td><td class="right">{{ fleet_pct }}%</td></tr>
  <tr class="total-row"><td>Total</td><td class="right">{{ total_fms }}</td><td class="right">100%</td></tr>
</table>

<h2>Shift-wise Summary</h2>
<table>
  <tr>
    <th>Shift</th><th>Attendant</th><th>Start</th><th>End</th>
    <th class="right">FMS (₹)</th><th class="right">Cash (₹)</th>
    <th class="right">Variance (₹)</th><th>Status</th>
  </tr>
  {% for s in shifts %}
  <tr>
    <td>{{ s.label }}</td>
    <td>{{ s.attendant }}</td>
    <td>{{ s.start }}</td>
    <td>{{ s.end }}</td>
    <td class="right">{{ s.fms }}</td>
    <td class="right">{{ s.cash }}</td>
    <td class="right {{ s.var_class }}">{{ s.variance }}</td>
    <td class="{{ s.var_class }}">{{ s.status }}</td>
  </tr>
  {% endfor %}
</table>

{% if attendant_shortage %}
<h2>Attendant Shortage Summary</h2>
<table>
  <tr><th>Attendant</th><th class="right">Total Shortage (₹)</th><th class="right">Shifts</th></tr>
  {% for a in attendant_shortage %}
  <tr>
    <td>{{ a.name }}</td>
    <td class="right shortage">{{ a.total_shortage }}</td>
    <td class="right">{{ a.shift_count }}</td>
  </tr>
  {% endfor %}
</table>
{% endif %}

{% if trend_rows %}
<h2>7-Day Trend</h2>
<table>
  <tr><th>Date</th><th class="right">FMS Total (₹)</th><th class="right">Net Variance (₹)</th><th>Status</th></tr>
  {% for t in trend_rows %}
  <tr>
    <td>{{ t.date }}</td>
    <td class="right">{{ t.fms }}</td>
    <td class="right {{ t.var_class }}">{{ t.variance }}</td>
    <td class="{{ t.var_class }}">{{ t.status }}</td>
  </tr>
  {% endfor %}
</table>
{% endif %}

<div class="footer">
  Generated: {{ generated_at }} IST &nbsp;|&nbsp; PetroLedger v1.0 &nbsp;|&nbsp; CONFIDENTIAL — OWNER ONLY
</div>
</body>
</html>"""


# ── Service ──────────────────────────────────────────────────────────────


class DailyReportService:
    """Generates daily summary PDF + Excel reports for a site."""

    def __init__(self) -> None:
        self._jinja_env = Environment(autoescape=select_autoescape(["html"]))
        self._template = self._jinja_env.from_string(_DAILY_TEMPLATE)

    async def generate_pdf(
        self,
        site_id: uuid.UUID,
        report_date: date,
        db: AsyncSession,
    ) -> Path:
        """Generate daily PDF report and return its path."""
        context = await self._build_context(site_id, report_date, db)
        html = self._template.render(**context)
        pdf_path = self._pdf_path(site_id, report_date)
        _write_pdf(html, pdf_path)
        log.info("daily_pdf_generated", site_id=str(site_id), date=str(report_date))
        return pdf_path

    async def generate_excel(
        self,
        site_id: uuid.UUID,
        report_date: date,
        db: AsyncSession,
    ) -> Path:
        """Generate daily Excel report and return its path."""
        context = await self._build_context(site_id, report_date, db)
        xl_path = self._excel_path(site_id, report_date)
        _write_excel(context, xl_path)
        log.info("daily_excel_generated", site_id=str(site_id), date=str(report_date))
        return xl_path

    # ------------------------------------------------------------------
    # Context building
    # ------------------------------------------------------------------

    async def _build_context(
        self,
        site_id: uuid.UUID,
        report_date: date,
        db: AsyncSession,
    ) -> dict[str, Any]:
        from app.utils.datetime import IST

        # Load site name
        pump_result = await db.execute(select(Pump).where(Pump.org_id == site_id).limit(1))
        pump = pump_result.scalar_one_or_none()
        site_name = str(site_id)  # fallback
        if pump and hasattr(pump, "organization") and pump.organization:
            site_name = pump.organization.name

        # Load all shifts for the day
        day_start = datetime(report_date.year, report_date.month, report_date.day, tzinfo=IST)
        day_end = day_start + timedelta(days=1)
        prev_22h = datetime(
            (report_date - timedelta(days=1)).year,
            (report_date - timedelta(days=1)).month,
            (report_date - timedelta(days=1)).day,
            22, 0, tzinfo=IST,
        )
        today_6h = datetime(report_date.year, report_date.month, report_date.day, 6, 0, tzinfo=IST)

        shifts_result = await db.execute(
            select(Shift).where(
                Shift.pump_id.in_(
                    select(Pump.id).where(Pump.org_id == site_id)
                ),
                or_(
                    (Shift.start_time >= day_start.astimezone(UTC))
                    & (Shift.start_time < day_end.astimezone(UTC)),
                    (Shift.start_time >= prev_22h.astimezone(UTC))
                    & (Shift.start_time < today_6h.astimezone(UTC)),
                ),
            )
        )
        shifts = list(shifts_result.scalars().all())
        shift_ids = [s.id for s in shifts]

        # Aggregate totals
        totals = await self._aggregate_totals(shift_ids, db)
        shift_rows = await self._build_shift_rows(shifts, db)
        attendant_shortage = self._build_attendant_shortage(shift_rows)
        trend_rows = await self._build_trend(site_id, report_date, db)

        # Payment mode percentages
        total_fms = totals["fms"]
        total_upi = totals["upi"]
        total_card = totals["card"]
        total_fleet = totals["fleet"]
        total_cash = totals["cash"]
        total_digital = total_upi + total_card + total_fleet

        def pct(part: Decimal, whole: Decimal) -> str:
            if whole == 0:
                return "0.0"
            return f"{(part / whole * 100):.1f}"

        net_variance = sum(
            (Decimal(r["variance_raw"]) for r in shift_rows), Decimal("0")
        )

        return {
            "site_name": site_name,
            "report_date": report_date.strftime("%d %b %Y"),
            "shift_count": len(shifts),
            "total_fms": f"{total_fms:,.2f}",
            "total_cash": f"{total_cash:,.2f}",
            "total_digital": f"{total_digital:,.2f}",
            "total_upi": f"{total_upi:,.2f}",
            "total_card": f"{total_card:,.2f}",
            "total_fleet": f"{total_fleet:,.2f}",
            "net_variance": net_variance,
            "cash_pct": pct(total_cash, total_fms),
            "upi_pct": pct(total_upi, total_fms),
            "card_pct": pct(total_card, total_fms),
            "fleet_pct": pct(total_fleet, total_fms),
            "shifts": shift_rows,
            "attendant_shortage": attendant_shortage,
            "trend_rows": trend_rows,
            "generated_at": to_ist(datetime.now(UTC)).strftime("%d/%m/%Y %H:%M"),
            # raw fields for Excel
            "_site_id": str(site_id),
            "_report_date_iso": str(report_date),
            "_totals": totals,
        }

    # ------------------------------------------------------------------
    # Data helpers
    # ------------------------------------------------------------------

    @staticmethod
    async def _aggregate_totals(
        shift_ids: list[uuid.UUID], db: AsyncSession
    ) -> dict[str, Decimal]:
        if not shift_ids:
            return {k: Decimal("0") for k in ["fms", "upi", "card", "fleet", "cash"]}

        async def _q(stmt):
            return Decimal(str((await db.execute(stmt)).scalar_one()))

        fms = await _q(
            select(func.coalesce(func.sum(FmsTransaction.amount), 0)).where(
                FmsTransaction.shift_id.in_(shift_ids),
                FmsTransaction.status == FmsTxnStatus.COMPLETED,
                FmsTransaction.is_deleted.is_(False),
            )
        )
        upi = await _q(
            select(func.coalesce(func.sum(UPITransaction.amount), 0)).where(
                UPITransaction.shift_id.in_(shift_ids)
            )
        )
        card = await _q(
            select(func.coalesce(func.sum(PosBatchSettlement.gross_amount), 0)).where(
                PosBatchSettlement.shift_id.in_(shift_ids),
                PosBatchSettlement.is_deleted.is_(False),
            )
        )
        fleet = await _q(
            select(func.coalesce(func.sum(FleetTransaction.total_amount), 0)).where(
                FleetTransaction.shift_id.in_(shift_ids),
                FleetTransaction.is_deleted.is_(False),
            )
        )
        cash = await _q(
            select(func.coalesce(func.sum(CashEntry.physical_cash), 0)).where(
                CashEntry.shift_id.in_(shift_ids),
                CashEntry.is_deleted.is_(False),
            )
        )
        return {"fms": fms, "upi": upi, "card": card, "fleet": fleet, "cash": cash}

    @staticmethod
    async def _build_shift_rows(
        shifts: list[Shift], db: AsyncSession
    ) -> list[dict[str, Any]]:
        rows = []
        for i, shift in enumerate(shifts):
            recon_result = await db.execute(
                select(ReconciliationResult).where(ReconciliationResult.shift_id == shift.id)
            )
            recon = recon_result.scalar_one_or_none()

            start_ist = to_ist(shift.start_time)
            end_ist = to_ist(shift.end_time) if shift.end_time else None

            fms_total = Decimal(str(
                (await db.execute(
                    select(func.coalesce(func.sum(FmsTransaction.amount), 0)).where(
                        FmsTransaction.shift_id == shift.id,
                        FmsTransaction.status == FmsTxnStatus.COMPLETED,
                        FmsTransaction.is_deleted.is_(False),
                    )
                )).scalar_one()
            ))
            cash_total = Decimal(str(
                (await db.execute(
                    select(func.coalesce(func.sum(CashEntry.physical_cash), 0)).where(
                        CashEntry.shift_id == shift.id,
                        CashEntry.is_deleted.is_(False),
                    )
                )).scalar_one()
            ))

            variance = recon.variance if recon else Decimal("0")
            if variance == 0:
                var_class, status = "match", "MATCH"
            elif variance > 0:
                var_class, status = "shortage", "SHORTAGE"
            else:
                var_class, status = "excess", "EXCESS"

            attendant = (
                shift.worker.user.email
                if hasattr(shift.worker, "user") and shift.worker.user
                else str(shift.worker_id)[:8]
            ) if hasattr(shift, "worker") and shift.worker else "—"

            rows.append({
                "label": f"Shift {i + 1}",
                "attendant": attendant,
                "start": start_ist.strftime("%H:%M"),
                "end": end_ist.strftime("%H:%M") if end_ist else "—",
                "fms": f"{fms_total:,.2f}",
                "cash": f"{cash_total:,.2f}",
                "variance": f"{abs(variance):,.2f}",
                "variance_raw": str(variance),
                "var_class": var_class,
                "status": status,
                "attendant_id": str(shift.worker_id),
                "shortage_amount": variance if variance > 0 else Decimal("0"),
            })
        return rows

    @staticmethod
    def _build_attendant_shortage(
        shift_rows: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        totals: dict[str, dict] = {}
        for row in shift_rows:
            att_id = row["attendant_id"]
            if att_id not in totals:
                totals[att_id] = {
                    "name": row["attendant"],
                    "total_shortage": Decimal("0"),
                    "shift_count": 0,
                }
            totals[att_id]["total_shortage"] += row["shortage_amount"]
            totals[att_id]["shift_count"] += 1

        return [
            {
                "name": v["name"],
                "total_shortage": f"{v['total_shortage']:,.2f}",
                "shift_count": v["shift_count"],
            }
            for v in totals.values()
            if v["total_shortage"] > 0
        ]

    @staticmethod
    async def _build_trend(
        site_id: uuid.UUID,
        report_date: date,
        db: AsyncSession,
    ) -> list[dict[str, Any]]:
        """Build 7-day trailing revenue + variance trend."""
        from app.utils.datetime import IST

        trend = []
        for delta in range(6, -1, -1):
            d = report_date - timedelta(days=delta)
            day_start = datetime(d.year, d.month, d.day, tzinfo=IST).astimezone(UTC)
            day_end = (day_start + timedelta(days=1))

            shifts_result = await db.execute(
                select(Shift.id).where(
                    Shift.pump_id.in_(select(Pump.id).where(Pump.org_id == site_id)),
                    Shift.start_time >= day_start,
                    Shift.start_time < day_end,
                )
            )
            sids = [r[0] for r in shifts_result.all()]
            if not sids:
                trend.append({
                    "date": d.strftime("%d %b"),
                    "fms": "—",
                    "variance": "—",
                    "var_class": "",
                    "status": "—",
                })
                continue

            fms_raw = (await db.execute(
                select(func.coalesce(func.sum(FmsTransaction.amount), 0)).where(
                    FmsTransaction.shift_id.in_(sids),
                    FmsTransaction.status == FmsTxnStatus.COMPLETED,
                    FmsTransaction.is_deleted.is_(False),
                )
            )).scalar_one()

            recon_rows = await db.execute(
                select(ReconciliationResult.variance).where(
                    ReconciliationResult.shift_id.in_(sids)
                )
            )
            variances = [Decimal(str(r[0])) for r in recon_rows.all()]
            net_var = sum(variances, Decimal("0"))

            if net_var == 0:
                var_class, status = "match", "MATCH"
            elif net_var > 0:
                var_class, status = "shortage", "SHORTAGE"
            else:
                var_class, status = "excess", "EXCESS"

            trend.append({
                "date": d.strftime("%d %b"),
                "fms": f"{Decimal(str(fms_raw)):,.2f}",
                "variance": f"{abs(net_var):,.2f}",
                "var_class": var_class,
                "status": status,
            })
        return trend

    # ------------------------------------------------------------------
    # File paths
    # ------------------------------------------------------------------

    @staticmethod
    def _pdf_path(site_id: uuid.UUID, report_date: date) -> Path:
        _REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        return _REPORTS_DIR / f"{site_id}_{report_date}_daily.pdf"

    @staticmethod
    def _excel_path(site_id: uuid.UUID, report_date: date) -> Path:
        _REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        return _REPORTS_DIR / f"{site_id}_{report_date}_daily.xlsx"


# ── PDF + Excel writers ──────────────────────────────────────────────────

def _write_pdf(html: str, output_path: Path) -> None:
    try:
        from weasyprint import HTML
        HTML(string=html).write_pdf(str(output_path))
    except ImportError:
        output_path.with_suffix(".html").write_text(html, encoding="utf-8")
        log.warning("weasyprint_not_installed", output=str(output_path))


def _write_excel(context: dict[str, Any], output_path: Path) -> None:
    """Write daily summary to an Excel workbook (openpyxl)."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        log.warning("openpyxl_not_installed", output=str(output_path))
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3A5C")

    def _header(ws, row_idx: int, values: list[str]) -> None:
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    ws.cell(1, 1, "PetroLedger Daily Summary").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Site: {context['site_name']}")
    ws.cell(3, 1, f"Date: {context['report_date']}")
    ws.cell(4, 1, f"Generated: {context['generated_at']} IST")

    _header(ws, 6, ["Metric", "Amount (₹)", "% of Revenue"])
    summary_rows = [
        ("Total FMS Revenue", context["total_fms"], "100%"),
        ("Cash Collected", context["total_cash"], f"{context['cash_pct']}%"),
        ("UPI Payments", context["total_upi"], f"{context['upi_pct']}%"),
        ("Card / POS", context["total_card"], f"{context['card_pct']}%"),
        ("Fleet Cards", context["total_fleet"], f"{context['fleet_pct']}%"),
    ]
    for i, (label, amount, pct_val) in enumerate(summary_rows, 7):
        ws.cell(i, 1, label)
        ws.cell(i, 2, amount)
        ws.cell(i, 3, pct_val)

    # ── Sheet 2: Shift Detail ────────────────────────────────────────────
    ws2 = wb.create_sheet("Shift Detail")
    _header(ws2, 1, ["Shift", "Attendant", "Start", "End", "FMS (₹)", "Cash (₹)", "Variance (₹)", "Status"])
    for i, s in enumerate(context["shifts"], 2):
        ws2.cell(i, 1, s["label"])
        ws2.cell(i, 2, s["attendant"])
        ws2.cell(i, 3, s["start"])
        ws2.cell(i, 4, s["end"])
        ws2.cell(i, 5, s["fms"])
        ws2.cell(i, 6, s["cash"])
        ws2.cell(i, 7, s["variance"])
        ws2.cell(i, 8, s["status"])

    # ── Sheet 3: 7-Day Trend ─────────────────────────────────────────────
    ws3 = wb.create_sheet("7-Day Trend")
    _header(ws3, 1, ["Date", "FMS Total (₹)", "Net Variance (₹)", "Status"])
    for i, t in enumerate(context["trend_rows"], 2):
        ws3.cell(i, 1, t["date"])
        ws3.cell(i, 2, t["fms"])
        ws3.cell(i, 3, t["variance"])
        ws3.cell(i, 4, t["status"])

    for ws_sheet in [ws, ws2, ws3]:
        for col in ws_sheet.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws_sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(str(output_path))
